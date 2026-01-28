"""Excel出力モジュール - フロー型レイアウト・罫線・書式設定

ビジュアルカラムの列構成:
  取引先 | 顧客 | 案件名 | 名前 | 部署 | グレード
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

import config
from src.processor import Division, Partner, Client, Project, Member

# --- 定数 ---
# A列は固定空列（左マージン）
LEFT_MARGIN = 1

# 1ビジュアルカラムあたりのExcel列数
EXCEL_COLS_PER_VISUAL = 7  # 取引先 | 顧客 | 案件名 | 名前 | 所属 | 空列 | グレード
GAP_COLS = 1
STRIDE = EXCEL_COLS_PER_VISUAL + GAP_COLS

# 列オフセット（0始まり）
COL_PARTNER = 0
COL_CLIENT = 1
COL_PROJECT = 2
COL_NAME = 3
COL_DEPT = 4
COL_EMPTY = 5
COL_GRADE = 6

# タイトル行 + 空行
HEADER_ROWS = 2
CONTENT_START_ROW = HEADER_ROWS + 1

# 罫線スタイル
MEDIUM = Side(style="medium")
THIN = Side(style="thin")
HAIR = Side(style="hair")

# 取引先ブロック間の空白行数
PARTNER_GAP_ROWS = 1

# 列幅変換パラメータ
# Excel は Normal style フォントの MaxDigitWidth(MDW) とピクセルパディングで
# XML格納幅 ⇔ UI表示幅を変換する。游ゴシック 11pt: MDW=9, padding=7。
_COL_WIDTH_MDW = 9
_COL_WIDTH_PADDING = 7


def _display_to_stored_width(display_width: float) -> float:
    """Excel UI 上の表示幅を XML 格納幅に変換する。

    stored = Truncate((display + padding/MDW) * 256) / 256
    """
    if display_width <= 0:
        return 0
    return int((display_width + _COL_WIDTH_PADDING / _COL_WIDTH_MDW) * 256) / 256


def _make_title_date() -> str:
    if config.CHART_TITLE_DATE:
        return config.CHART_TITLE_DATE
    now = datetime.now()
    reiwa_year = now.year - 2018
    return f"R{reiwa_year}年{now.month}月"


def _make_font(font_cfg: dict) -> Font:
    return Font(**font_cfg)


class FlowLayout:
    """フロー型カラムレイアウトエンジン。"""

    def __init__(self, ws):
        self.ws = ws
        self.cols_per_page = config.LAYOUT["columns_per_page"]
        self.max_rows = config.LAYOUT["max_rows_per_column"]

        self.visual_col = 0
        self.page = 0
        self.row = CONTENT_START_ROW

        # ページ段管理（5カラムを超えたら下段＝次の印刷ページへ）
        self.page_start_row = CONTENT_START_ROW
        self.page_max_row = CONTENT_START_ROW
        self.page_break_rows = []

    @property
    def _col_base(self) -> int:
        """現在のビジュアルカラムの開始Excel列番号（1始まり）。A列はマージン。"""
        return self.visual_col * STRIDE + 1 + LEFT_MARGIN

    def _col(self, offset: int) -> int:
        """列オフセットからExcel列番号を返す。"""
        return self._col_base + offset

    def can_fit(self, height: int) -> bool:
        return (self.row + height - 1) <= (self.page_start_row + self.max_rows - 1)

    def next_column(self):
        self.page_max_row = max(self.page_max_row, self.row)
        self.visual_col += 1
        self.row = self.page_start_row
        if self.visual_col >= self.cols_per_page:
            self.page += 1
            self.visual_col = 0
            # 現ページ最大行の直後にページブレークを挿入
            break_row = self.page_max_row
            self.page_break_rows.append(break_row)
            # 上部マージン（HEADER_ROWS分）を空けてコンテンツ開始
            self.page_start_row = break_row + HEADER_ROWS
            self.row = self.page_start_row
            self.page_max_row = self.page_start_row

    def ensure_fit(self, height: int):
        if not self.can_fit(height):
            self.next_column()

    def write_partner_clients(self, partner: Partner, division_key: str = ""):
        """取引先ブロックを顧客単位でカラムをまたいで書き込む。

        顧客ブロック（顧客名＋案件＋メンバー）は途中で切らない。
        カラムに収まらない場合は次カラムへ移動し、取引先名を再出力する。

        レイアウト:
          取引先名        ← COL_PARTNER   | 営業名    ← COL_GRADE
          （空行）                         | 〇名      ← COL_GRADE
            顧客名        ← COL_CLIENT
            （空行）
              案件名      ← COL_PROJECT
              名前|部署|グレード  ← COL_NAME〜COL_GRADE
              ...
        """
        PARTNER_HEADER_ROWS = 2

        font_partner = _make_font(config.LAYOUT["font_partner"])
        font_client = _make_font(config.LAYOUT["font_client"])
        font_project = _make_font(config.LAYOUT["font_project"])
        font_person = _make_font(config.LAYOUT["font_person"])
        font_info = Font(name="ＭＳ Ｐゴシック", size=9, italic=True)
        align_right = Alignment(horizontal="right")

        # --- セグメント（カラム内の取引先部分ブロック）管理 ---
        seg = {}

        def _begin_segment():
            """新しいセグメントを開始し、取引先ヘッダーを書き込む。"""
            seg.clear()
            seg["start_row"] = self.row
            seg["col_base"] = self._col_base
            seg["partner_row"] = self.row
            seg["member_ranges"] = []
            seg["border_starts"] = {}
            seg["merged_rows"] = set()
            seg["client_vline_ranges"] = []
            seg["project_vline_ranges"] = []
            seg["first_client"] = True

            # 取引先名（右5列 + 下1行をセル結合）
            pr = self.row
            cell = self.ws.cell(
                row=pr, column=self._col(COL_PARTNER),
                value=partner.display_name,
            )
            cell.font = font_partner
            cell.alignment = Alignment(vertical="center", shrink_to_fit=True)
            self.ws.merge_cells(
                start_row=pr, start_column=self._col(COL_PARTNER),
                end_row=pr + 1, end_column=self._col(COL_PARTNER + 5),
            )
            seg["merged_rows"].add(pr)

            # 担当営業名（上段 COL_GRADE）
            cell_div = self.ws.cell(
                row=pr, column=self._col(COL_GRADE),
                value=f"営業:{division_key}",
            )
            cell_div.font = font_info
            cell_div.alignment = align_right

            # 人数合計（下段 COL_GRADE）
            cell_cnt = self.ws.cell(
                row=pr + 1, column=self._col(COL_GRADE),
                value=f"{partner.count}名",
            )
            cell_cnt.font = font_info
            cell_cnt.alignment = align_right

            self.row += PARTNER_HEADER_ROWS
            seg["border_starts"][self.row] = COL_CLIENT

        def _end_segment():
            """現セグメントの罫線を適用して閉じる。"""
            if not seg:
                return
            end_row = self.row - 1
            if end_row < seg["start_row"]:
                return
            col_base = seg["col_base"]
            col_end = col_base + EXCEL_COLS_PER_VISUAL - 1
            _apply_partner_borders(
                self.ws, seg["start_row"], end_row, col_base, col_end,
                seg["member_ranges"], seg["border_starts"],
                seg["merged_rows"],
                partner_row=seg["partner_row"],
                client_vline_ranges=seg["client_vline_ranges"],
                project_vline_ranges=seg["project_vline_ranges"],
            )

        # --- 最初のセグメント開始 ---
        first_h = partner.clients[0].row_height() if partner.clients else 0
        self.ensure_fit(PARTNER_HEADER_ROWS + first_h)
        _begin_segment()

        # --- 顧客ブロックを順に書き込み ---
        for client in partner.clients:
            client_h = client.row_height()

            # 現カラムに収まらなければセグメントを閉じて改カラム
            if not self.can_fit(client_h):
                _end_segment()
                self.next_column()
                self.ensure_fit(PARTNER_HEADER_ROWS + client_h)
                _begin_segment()

            if not seg["first_client"]:
                seg["border_starts"][self.row] = COL_CLIENT
            seg["first_client"] = False

            # 顧客名（col6まで + 下1行をセル結合）
            client_row = self.row
            cell_c = self.ws.cell(
                row=client_row, column=self._col(COL_CLIENT),
                value=client.name,
            )
            cell_c.font = font_client
            cell_c.alignment = Alignment(vertical="center", shrink_to_fit=True)
            self.ws.merge_cells(
                start_row=client_row, start_column=self._col(COL_CLIENT),
                end_row=client_row + 1, end_column=self._col(COL_EMPTY),
            )
            seg["merged_rows"].add(client_row)

            # 顧客人数合計（下段 COL_GRADE）
            cell_cc = self.ws.cell(
                row=client_row + 1, column=self._col(COL_GRADE),
                value=f"{client.count}名",
            )
            cell_cc.font = font_info
            cell_cc.alignment = align_right

            self.row += 2  # 顧客名行 + 空行（結合に含まれる）
            seg["border_starts"][self.row] = COL_PROJECT

            for pi, project in enumerate(client.projects):
                if pi > 0:
                    seg["border_starts"][self.row] = COL_PROJECT

                # 案件名（col6までセル結合）
                proj_row = self.row
                cell_p = self.ws.cell(
                    row=proj_row, column=self._col(COL_PROJECT),
                    value=project.name,
                )
                cell_p.font = font_project
                cell_p.alignment = Alignment(shrink_to_fit=True)
                self.ws.merge_cells(
                    start_row=proj_row, start_column=self._col(COL_PROJECT),
                    end_row=proj_row, end_column=self._col(COL_EMPTY),
                )

                # 案件人数（COL_GRADE）
                cell_pc = self.ws.cell(
                    row=proj_row, column=self._col(COL_GRADE),
                    value=f"{project.count}名",
                )
                cell_pc.font = font_info
                cell_pc.alignment = align_right

                self.row += 1  # 案件名行
                seg["border_starts"][self.row] = COL_NAME

                # メンバー（案件名の直下）
                first_member_row = self.row
                for member in project.members:
                    r = self.row
                    cell_n = self.ws.cell(
                        row=r, column=self._col(COL_NAME),
                        value=member.name,
                    )
                    cell_n.font = font_person
                    cell_n.alignment = Alignment(shrink_to_fit=True)
                    self.ws.cell(
                        row=r, column=self._col(COL_DEPT),
                        value=member.dept,
                    ).font = font_person
                    if not member.is_bp and member.grade:
                        self.ws.cell(
                            row=r, column=self._col(COL_GRADE),
                            value=member.grade,
                        ).font = font_person
                    self.row += 1
                last_member_row = self.row - 1
                if project.count >= 2:
                    seg["member_ranges"].append(
                        (first_member_row, last_member_row))
                    for mr in range(first_member_row + 1, last_member_row + 1):
                        seg["border_starts"][mr] = COL_NAME

                # col3 右縦罫線: 案件名の1行下からメンバー最終行まで
                if project.count >= 1:
                    seg["project_vline_ranges"].append(
                        (proj_row + 1, last_member_row))

            # col2 右縦罫線: 顧客名セルの2行下から顧客最終行まで
            seg["client_vline_ranges"].append(
                (client_row + 2, self.row - 1))

        # 最終セグメントを閉じる
        _end_segment()

        # 取引先ブロック間の空白行
        self.row += PARTNER_GAP_ROWS


def _is_member_pair(row_a: int, row_b: int, member_ranges: list) -> bool:
    """row_a と row_b が同一案件のメンバー行同士かを判定する。"""
    for first, last in member_ranges:
        if first <= row_a <= last and first <= row_b <= last:
            return True
    return False


def _resolve_h_border(row_boundary: int, col_offset: int,
                      member_ranges: list, border_starts: dict) -> Side:
    """内部の横罫線スタイルを返す。

    row_boundary: 罫線が入る行（上の行と下の行の境界 = 下の行番号）
    col_offset:   現在の列オフセット（0始まり）
    """
    if row_boundary not in border_starts:
        return Side()  # 罫線なし

    start_col = border_starts[row_boundary]
    if col_offset < start_col:
        return Side()  # この列まで罫線は届かない

    # メンバー間は hair、それ以外は thin
    if _is_member_pair(row_boundary - 1, row_boundary, member_ranges):
        return HAIR
    return THIN


def _apply_partner_borders(ws, row_start, row_end, col_start, col_end,
                           member_ranges: list, border_starts: dict,
                           merged_rows: set, partner_row: int = 0,
                           client_vline_ranges: list = None,
                           project_vline_ranges: list = None):
    """取引先ブロックに罫線を適用する。

    - 外枠: medium（太線）
    - 内部横罫線: 階層レベルに応じて開始列が異なる
      - 顧客境界: COL_CLIENT (col2) から
      - 案件境界: COL_PROJECT (col3) から
      - メンバー間: COL_NAME (col4) から、hair
    - 結合セル内部: 罫線なし
    - col1右辺: thin（取引先名セルの2行下から）
    - col2右辺: thin（各顧客名セルの2行下から顧客最終行まで）
    - col3右辺: thin（各案件名セルの1行下からメンバー最終行まで）
    """
    if client_vline_ranges is None:
        client_vline_ranges = []
    if project_vline_ranges is None:
        project_vline_ranges = []
    col_base = col_start

    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            col_offset = col - col_base

            # --- 上辺 ---
            if row == row_start:
                top = MEDIUM
            elif (row - 1) in merged_rows:
                top = Side()  # 結合セル内部
            else:
                top = _resolve_h_border(
                    row, col_offset, member_ranges, border_starts)

            # --- 下辺 ---
            if row == row_end:
                bottom = MEDIUM
            elif row in merged_rows:
                bottom = Side()  # 結合セル内部
            else:
                bottom = _resolve_h_border(
                    row + 1, col_offset, member_ranges, border_starts)

            # --- 左辺・右辺 ---
            left = MEDIUM if col == col_start else Side()
            right = MEDIUM if col == col_end else Side()

            # col1 右辺: 取引先名の2行下から thin
            if col_offset == COL_PARTNER and row >= partner_row + 2:
                right = THIN

            # col2 右辺: 各顧客名の2行下から顧客最終行まで thin
            if col_offset == COL_CLIENT:
                for cv_start, cv_end in client_vline_ranges:
                    if cv_start <= row <= cv_end:
                        right = THIN
                        break

            # col3 右辺: 各案件名の1行下からメンバー最終行まで thin
            if col_offset == COL_PROJECT:
                for pv_start, pv_end in project_vline_ranges:
                    if pv_start <= row <= pv_end:
                        right = THIN
                        break

            ws.cell(row=row, column=col).border = Border(
                top=top, bottom=bottom, left=left, right=right,
            )


def _setup_column_widths(ws, total_visual_cols: int):
    """全ビジュアルカラムの列幅を設定する。

    config の値は Excel UI 上の表示幅なので、XML 格納幅に変換して設定する。
    """
    layout = config.LAYOUT
    _w = _display_to_stored_width

    # A列（左マージン空列）
    ws.column_dimensions["A"].width = _w(layout["col_width_margin"])

    for vc in range(total_visual_cols):
        base = vc * STRIDE + 1 + LEFT_MARGIN
        ws.column_dimensions[get_column_letter(base + COL_PARTNER)].width = _w(layout["col_width_partner"])
        ws.column_dimensions[get_column_letter(base + COL_CLIENT)].width = _w(layout["col_width_client"])
        ws.column_dimensions[get_column_letter(base + COL_PROJECT)].width = _w(layout["col_width_project"])
        ws.column_dimensions[get_column_letter(base + COL_NAME)].width = _w(layout["col_width_name"])
        ws.column_dimensions[get_column_letter(base + COL_DEPT)].width = _w(layout["col_width_dept"])
        ws.column_dimensions[get_column_letter(base + COL_EMPTY)].width = _w(layout["col_width_empty"])
        ws.column_dimensions[get_column_letter(base + COL_GRADE)].width = _w(layout["col_width_grade"])
        if GAP_COLS > 0:
            ws.column_dimensions[get_column_letter(base + EXCEL_COLS_PER_VISUAL)].width = _w(layout["col_width_gap"])


def _setup_print(ws, page_break_rows: list, last_row: int):
    """印刷設定（A4横、印刷範囲固定、行ページブレーク）。"""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.scale = 48

    # 余白（cm → inches）
    _CM = 1 / 2.54
    ws.page_margins.top = 0.5 * _CM
    ws.page_margins.bottom = 0.5 * _CM
    ws.page_margins.left = 0.8 * _CM
    ws.page_margins.right = 0.8 * _CM
    ws.page_margins.header = 0.8 * _CM
    ws.page_margins.footer = 0.8 * _CM

    # 印刷範囲: A列〜AO列（5カラム分）を1ページ幅として固定
    last_col_num = LEFT_MARGIN + config.LAYOUT["columns_per_page"] * STRIDE
    last_col = get_column_letter(last_col_num)
    ws.print_area = f"A1:{last_col}{last_row}"

    # 列方向のページ境界: AO列の次列で区切り、横幅を1ページに固定
    ws.col_breaks.append(Break(id=last_col_num + 1))

    for break_row in page_break_rows:
        ws.row_breaks.append(Break(id=break_row))


def generate(divisions: list[Division], output_dir: Path) -> Path:
    """構成図Excelを生成する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "構成図"

    # Normal style のフォントをＭＳ Ｐゴシックに設定
    # （列幅の解釈に使われる MaxDigitWidth がフォントに依存するため）
    for ns in wb._named_styles:
        if ns.name == "Normal":
            ns.font = Font(name="ＭＳ Ｐゴシック", size=11)
            break

    # 目盛線を非表示、改ページプレビューで開く
    ws.sheet_view.showGridLines = False
    ws.sheet_view.view = "pageBreakPreview"

    # タイトル行（1行目は空、2行目にタイトル）
    title = f"【{_make_title_date()}】"
    ws.cell(row=2, column=1 + LEFT_MARGIN, value=title).font = _make_font(
        config.LAYOUT["font_title"]
    )

    # フローレイアウトで書き込み
    flow = FlowLayout(ws)

    for division in divisions:
        for partner in division.partners:
            flow.write_partner_clients(partner, division_key=division.key)

    # 列幅・印刷設定（常に5カラム分の列幅を固定）
    total_visual_cols = flow.cols_per_page
    _setup_column_widths(ws, total_visual_cols)
    last_row = max(flow.page_max_row, flow.row)
    _setup_print(ws, flow.page_break_rows, last_row)

    # 出力
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"構成図_{date_str}.xlsx"
    wb.save(output_path)

    return output_path
